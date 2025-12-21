# routers/conversores.py
import openpyxl
import io
from typing import Optional
from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from starlette.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from unidecode import unidecode

from .. import crud, models
from ..database import engine, AsyncSession
from ..auth import get_current_active_user
from ..config import settings
from pathlib import Path

# Resolve resource paths relative to this file so the code works
# when run inside Docker (where the working directory may be different).
PACKAGE_ROOT = Path(__file__).resolve().parent.parent
RESOURCES_BASE = PACKAGE_ROOT / "resources" / "base"


def find_resource_file(filename: str) -> str:
    """Return the first existing candidate path for a resource file.

    Candidates (in order):
      - package-root/resources/base/<filename>  (common when resources are inside the app folder)
      - package-root.parent/resources/base/<filename>  (common when resources are sibling to app/)
      - Path.cwd()/resources/base/<filename>  (fallback to current working directory)

    If none exist, return the first candidate (so the existing FileNotFoundError shows a sensible path).
    """
    candidates = [
        RESOURCES_BASE / filename,
        (PACKAGE_ROOT.parent / "resources" / "base" / filename),
        (Path.cwd() / "resources" / "base" / filename),
    ]

    for cand in candidates:
        try:
            if cand.exists():
                return str(cand)
        except Exception:
            # If any permission/IO oddity happens, ignore and try next
            continue

    # none found; return the first candidate for clearer error reporting later
    return str(candidates[0])

router = APIRouter(
    tags=["Conversores"],
    dependencies=[Depends(get_current_active_user)]
)

REMUNERACAO_BASE_FILE_PATH = find_resource_file("Valid-Remuneracao.xlsx")
TECD_BASE_FILE_PATH = find_resource_file("Valid-Tec-D.xlsx")
TECD_PRECO_FIXO = 17


def normalize_header(header: str) -> str:
    """Normaliza header removendo acentos, convertendo para lowercase, 
    tratando underscores e espaços como equivalentes."""
    normalized = unidecode(str(header)).lower().strip()
    normalized = normalized.replace("_", " ")
    while "  " in normalized:
        normalized = normalized.replace("  ", " ")
    return normalized.strip()


class RemuneracaoColumnInfo:
    def __init__(self, name: str, alternative_names: list[str]):
        self.name = name
        self.alternative_names = alternative_names

remuneracao_columns = [
    RemuneracaoColumnInfo("PEDIDO", []),
    RemuneracaoColumnInfo("SOLICITACAO", []),
    RemuneracaoColumnInfo("SKU", []),
    RemuneracaoColumnInfo("PRODUTO", []),
    RemuneracaoColumnInfo("TITULAR", []),
    RemuneracaoColumnInfo("CPF", []),
    RemuneracaoColumnInfo("EMAIL", []),
    RemuneracaoColumnInfo("CNPJ", []),
    RemuneracaoColumnInfo("RAZAO_SOCIAL", []),
    RemuneracaoColumnInfo("PONTO_CODIGO", ["Código Localidade"]),
    RemuneracaoColumnInfo("PONTO_ATENDIMENTO", ["Localidade Atendimento"]),
    RemuneracaoColumnInfo("NOME_AGENTE", ["Nome Agente Validação"]),
    RemuneracaoColumnInfo("CPF_AGENTE", ["CPF Agente Validação"]),
    RemuneracaoColumnInfo("DATA_1A_VERIFICACAO", ["Data Primeira Verificação"]),
    RemuneracaoColumnInfo("CODIGO_ORIGEM", []),
    RemuneracaoColumnInfo("NOME_ORIGEM", []),
    RemuneracaoColumnInfo("SERIAL_CERTIFICADO", []),
    RemuneracaoColumnInfo("DATA_EMISSAO", []),
    RemuneracaoColumnInfo("DATA_EXPIRACAO", []),
    RemuneracaoColumnInfo("DATA_REVOGACAO", []),
    RemuneracaoColumnInfo("DATA_HORA_1A_VERIFICACAO", ["Data Hora Primeira Verificação"]),
    RemuneracaoColumnInfo("DATA_ULTIMO_STATUS", []),
    RemuneracaoColumnInfo("ULTIMO_STATUS", []),
    RemuneracaoColumnInfo("CIDADE", ["Cidade Validação"]),
    RemuneracaoColumnInfo("ESTADO", ["Estado Validação"]),
    RemuneracaoColumnInfo("AR", []),
    RemuneracaoColumnInfo("TICKET_ANTERIOR", []),
    RemuneracaoColumnInfo("PRECO", []),
    RemuneracaoColumnInfo("DATA_VENDA", ["Data de Venda"]),
    RemuneracaoColumnInfo("PRECO BASE", []),
    RemuneracaoColumnInfo("% REM", []),
    RemuneracaoColumnInfo("R$ BIO", ["BIO"]),
    RemuneracaoColumnInfo("TOTAL R$", []),
    RemuneracaoColumnInfo("VOUCHER", []),
]


def is_remuneracao_name_matching(col_info: RemuneracaoColumnInfo, header_name: str) -> bool:
    """Verifica se um header corresponde a uma coluna de remuneração, incluindo nomes alternativos."""
    normalized_header = normalize_header(header_name)
    normalized_col_name = normalize_header(col_info.name)
    
    if normalized_header == normalized_col_name:
        return True
    
    for alt_name in col_info.alternative_names:
        normalized_alt = normalize_header(alt_name)
        if normalized_header == normalized_alt:
            return True
    
    return False


def find_data_column_for_base_column(
    base_col_name: str,
    col_info: Optional[RemuneracaoColumnInfo],
    data_header_map: dict
) -> int:
    """Encontra a coluna correspondente nos dados para uma coluna base."""
    if col_info:
        # Usa matching com nomes alternativos
        for data_header, data_col in data_header_map.items():
            if is_remuneracao_name_matching(col_info, data_header):
                return data_col
    else:
        # Usa matching simples por normalização
        normalized_base = normalize_header(base_col_name)
        for data_header, data_col in data_header_map.items():
            if normalize_header(data_header) == normalized_base:
                return data_col
    
    raise HTTPException(
        status_code=400,
        detail=f"Coluna '{base_col_name}' não encontrada na planilha enviada."
    )


def find_emissoes_sheet(wb_data):
    """Encontra a sheet de emissões no workbook."""
    for sheet_name in wb_data.sheetnames:
        if unidecode(sheet_name).lower().strip().startswith("emissoes"):
            return wb_data[sheet_name]
    return wb_data.active


@router.post(
    "/converter-remuneracao/",
    summary="Converte planilha de remuneração",
    description="Recebe uma planilha de dados (ex: Digiforte), valida suas colunas contra um template base, "
                "copia os dados e retorna o arquivo convertido."
)
async def converter_remuneracao(
    data_file: UploadFile = File(..., description="Planilha de dados a ser processada (ex: Digiforte.xlsx)")
):
    try:
        # 1. Carrega template base
        try:
            wb_base = openpyxl.load_workbook(REMUNERACAO_BASE_FILE_PATH)
            ws_base = wb_base.active
        except FileNotFoundError:
            raise HTTPException(
                status_code=500,
                detail="Erro interno no servidor: O arquivo de template base não foi encontrado."
            )

        # 2. Carrega arquivo de dados
        wb_data = openpyxl.load_workbook(data_file.file)
        ws_data = find_emissoes_sheet(wb_data)

        # 3. Prepara mapeamentos
        base_col_name_to_info = {col_info.name: col_info for col_info in remuneracao_columns}
        data_header_map = {cell.value: cell.column for cell in ws_data[1] if cell.value}
        base_header_ordered = [cell.value for cell in ws_base[1] if cell.value]

        # 4. Mapeia colunas base -> dados
        base_to_data_column_map = {}
        for base_col_name in base_header_ordered:
            if base_col_name == "VOUCHER":
                continue
            col_info = base_col_name_to_info.get(base_col_name)
            base_to_data_column_map[base_col_name] = find_data_column_for_base_column(
                base_col_name, col_info, data_header_map
            )

        # 5. Limpa template e copia dados
        if ws_base.max_row > 1:
            ws_base.delete_rows(idx=2, amount=ws_base.max_row - 1)

        for data_row in ws_data.iter_rows(min_row=2):
            if all(cell.value is None for cell in data_row):
                continue

            new_row_values = []
            for col_name in base_header_ordered:
                if col_name == "VOUCHER":
                    new_row_values.append("")
                else:
                    data_col_index = base_to_data_column_map[col_name]
                    new_row_values.append(data_row[data_col_index - 1].value)
            ws_base.append(new_row_values)

        # 6. Retorna arquivo convertido
        output_buffer = io.BytesIO()
        wb_base.save(output_buffer)
        output_buffer.seek(0)

        return StreamingResponse(
            output_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=Remuneracao-Convertida.xlsx"}
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"Ocorreu um erro inesperado: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Ocorreu um erro inesperado ao processar o arquivo: {str(e)}"
        )
    

class TecdColumnInfo:
    def __init__(self, name: str, index: int, alternative_names: list[str]):
        self.name = name
        self.index = index
        self.alternative_names = alternative_names

tecd_columns = [
    TecdColumnInfo("Pedido", 1, []),
    TecdColumnInfo("Solitação", 2, ["solicitacao"]),
    TecdColumnInfo("SKU", 3, []),
    TecdColumnInfo("Produto", 4, []),
    TecdColumnInfo("Titular", 5, []),
    TecdColumnInfo("CPF", 6, []),
    TecdColumnInfo("Email", 7, []),
    TecdColumnInfo("CNPJ", 8, []),
    TecdColumnInfo("Razão Social", 9, []),
    TecdColumnInfo("Codigo Localidade", 10, []),
    TecdColumnInfo("Localidade Atendimento", 11, []),
    TecdColumnInfo("Nome Agente Validação", 12, []),
    TecdColumnInfo("CPF Agente Validação", 13, []),
    TecdColumnInfo("Data Primeira Verificação", 14, []),
    TecdColumnInfo("Código Origem", 15, []),
    TecdColumnInfo("Nome Origem", 16, []),
    TecdColumnInfo("Preço", 17, []),
    TecdColumnInfo("Serial Certificado", 18, []),
    TecdColumnInfo("Data Emissão", 19, []),
    TecdColumnInfo("Data Expiração", 20, []),
    TecdColumnInfo("Data Revogação", 21, []),
    TecdColumnInfo("Data Hora Primeira Verificação", 22, []),
    TecdColumnInfo("Data Último Status", 23, []),
    TecdColumnInfo("Último Status", 24, []),
    TecdColumnInfo("Cidade", 25, ["cidade validacao"]),
    TecdColumnInfo("Estado", 26, ["estado validacao"]),
    TecdColumnInfo("AR", 27, []),
    TecdColumnInfo("TEC-D R$ ", 28, ["preco tec-d"]),
    TecdColumnInfo("TEC-D BIO ", 29, ["bio"]),
    TecdColumnInfo("TOTAL TEC-D R$", 30, ["total tec-d"]),
]

def is_name_matching(tecd_column_info: TecdColumnInfo, header_name: str) -> bool:
    if unidecode(header_name).lower().strip() == unidecode(tecd_column_info.name).lower().strip():
        return True
    for alt_name in tecd_column_info.alternative_names:
        if unidecode(header_name).lower().strip() == unidecode(alt_name).lower().strip():
            return True
    return False


@router.post(
    "/converter-tecd/",
    summary="Converte planilha de TEC-D",
    description="Recebe uma planilha de dados, valida suas colunas contra um template base, "
                "copia os dados e retorna o arquivo convertido."
)
async def converter_tecd(
    data_file: UploadFile = File(..., description="Planilha de dados a ser processada (ex: Digiforte.xlsx)")
):
    try:
        # 1. Carrega o "template" (base)
        try:
            wb_base = openpyxl.load_workbook(TECD_BASE_FILE_PATH)
            ws_base = wb_base.active
        except FileNotFoundError:
            print(f"Erro Crítico: Arquivo base não encontrado em: {TECD_BASE_FILE_PATH}")
            raise HTTPException(
                status_code=500, 
                detail="Erro interno no servidor: O arquivo de template base não foi encontrado."
            )

        # 2. Carrega o "arquivo de dados"
        wb_data = openpyxl.load_workbook(data_file.file)
        ws_data = wb_data.active
        for sheet_name in wb_data.sheetnames:
            if unidecode(sheet_name).lower().strip() == "emissoes":
                ws_data = wb_data[sheet_name]
                break
        
        # 3. Mapeia os cabeçalhos
        column_map = {}
        for col_info in tecd_columns:
            found = False
            for cell in ws_data[1]:
                if cell.value is None:
                    continue
                if is_name_matching(col_info, cell.value):
                    column_map[col_info.index] = cell.column
                    found = True
                    break
            if not found:
                detail_msg = f"Coluna '{col_info.name}' não encontrada na planilha enviada."
                print(f"Erro de Validação: {detail_msg}")
                print(f"Colunas disponíveis: {[cell.value for cell in ws_data[1] if cell.value]}")
                raise HTTPException(status_code=400, detail=detail_msg)

        # 5. Limpa o "template"
        if ws_base.max_row > 1:
            ws_base.delete_rows(idx=2, amount=ws_base.max_row - 1)

        # 6. Copia os dados
        for data_row in ws_data.iter_rows(min_row=2):
            if all(cell.value is None for cell in data_row):
                continue
            
            new_row_values = []
            for tecd_col in tecd_columns:
                data_col_index = column_map[tecd_col.index]
                value_to_copy = data_row[data_col_index - 1].value
                new_row_values.append(value_to_copy)
            
            # Lógica específica do TEC-D
            nome_agente = new_row_values[11].strip()
            cpf_agente = new_row_values[12].replace(".", "").replace("-", "").strip()
            nome_localidade = new_row_values[10].strip()
            is_virtual = bool(nome_localidade and "VIRTUAL" in nome_localidade.upper())

            if is_virtual:
                async with AsyncSession(engine) as session:
                    db_agente = await crud.get_agente_by_cpf(session, cpf=cpf_agente)
                    if not db_agente:
                        raise HTTPException(status_code=404, detail=f"Agente {nome_agente} com CPF {cpf_agente} não encontrado.")
                    
                    db_localidade = await crud.get_localidade(session, localidade_id=db_agente.localidade_id)
                    if not db_localidade:
                        raise HTTPException(status_code=404, detail=f"Agente {nome_agente} não associado a uma localidade física.")

                    new_row_values[9] = db_localidade.codigo_localidade
                    new_row_values[10] = db_localidade.nome

            new_row_values[27] = TECD_PRECO_FIXO
            new_row_values[29] = TECD_PRECO_FIXO

            ws_base.append(new_row_values)
        
        # 7. Salva em buffer
        output_buffer = io.BytesIO()
        wb_base.save(output_buffer)
        output_buffer.seek(0)

        # 8. Retorna o buffer
        return StreamingResponse(
            output_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=Remuneracao-Convertida.xlsx"}
        )

    except HTTPException as e:
        raise e
    except Exception as e:
        print(f"Ocorreu um erro inesperado: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500, 
            detail=f"Ocorreu um erro inesperado ao processar o arquivo: {str(e)}"
        )