# routers/remuneracao.py
import openpyxl
import io
from typing import Optional
from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from starlette.responses import StreamingResponse
from unidecode import unidecode

from ..auth import get_current_active_user
from .utils import find_resource_file

router = APIRouter(
    tags=["Conversores"],
    dependencies=[Depends(get_current_active_user)]
)

REMUNERACAO_BASE_FILE_PATH = find_resource_file("Valid-Remuneracao.xlsx")


def normalize_header(header: str) -> str:
    """Normaliza header removendo acentos, convertendo para lowercase, 
    tratando underscores e espaços como equivalentes."""
    normalized = unidecode(str(header)).lower().strip()
    normalized = normalized.replace("_", " ")
    while "  " in normalized:
        normalized = normalized.replace("  ", " ")
    return normalized.strip()


def find_relevant_sheets(wb_data):
    """Encontra todas as sheets relevantes (EMISSÕES e EMISSÃO AC)."""
    relevant_sheets = []
    for sheet_name in wb_data.sheetnames:
        normalized_name = unidecode(sheet_name).lower().strip()
        if normalized_name.startswith("emissoes") or normalized_name.startswith("emissao ac"):
            relevant_sheets.append(wb_data[sheet_name])
    
    # Se não encontrou nenhuma, usa a sheet ativa
    if not relevant_sheets:
        relevant_sheets.append(wb_data.active)
    
    return relevant_sheets


def process_sheet_data(ws_data, base_header_ordered, base_to_data_column_map):
    """Processa dados de uma sheet e retorna lista de linhas."""
    rows = []
    for data_row in ws_data.iter_rows(min_row=2):
        if all(cell.value is None for cell in data_row):
            continue

        new_row_values = []
        for col_name in base_header_ordered:
            if col_name == "VOUCHER":
                new_row_values.append("")
            else:
                data_col_index = base_to_data_column_map[col_name]
                new_row_values.append(data_row[data_col_index - 1].value)
        rows.append(new_row_values)
    return rows


class RemuneracaoColumnInfo:
    def __init__(self, name: str, alternative_names: list[str]):
        self.name = name
        self.alternative_names = alternative_names


remuneracao_columns = [
    RemuneracaoColumnInfo("PEDIDO", []),
    RemuneracaoColumnInfo("SOLICITACAO", []),
    RemuneracaoColumnInfo("SKU", []),
    RemuneracaoColumnInfo("PRODUTO", []),
    RemuneracaoColumnInfo("TITULAR", []),
    RemuneracaoColumnInfo("CPF", []),
    RemuneracaoColumnInfo("EMAIL", []),
    RemuneracaoColumnInfo("CNPJ", []),
    RemuneracaoColumnInfo("RAZAO_SOCIAL", []),
    RemuneracaoColumnInfo("PONTO_CODIGO", ["Código Localidade"]),
    RemuneracaoColumnInfo("PONTO_ATENDIMENTO", ["Localidade Atendimento"]),
    RemuneracaoColumnInfo("NOME_AGENTE", ["Nome Agente Validação"]),
    RemuneracaoColumnInfo("CPF_AGENTE", ["CPF Agente Validação"]),
    RemuneracaoColumnInfo("DATA_1A_VERIFICACAO", ["Data Primeira Verificação"]),
    RemuneracaoColumnInfo("CODIGO_ORIGEM", []),
    RemuneracaoColumnInfo("NOME_ORIGEM", []),
    RemuneracaoColumnInfo("SERIAL_CERTIFICADO", []),
    RemuneracaoColumnInfo("DATA_EMISSAO", []),
    RemuneracaoColumnInfo("DATA_EXPIRACAO", []),
    RemuneracaoColumnInfo("DATA_REVOGACAO", []),
    RemuneracaoColumnInfo("DATA_HORA_1A_VERIFICACAO", ["Data Hora Primeira Verificação"]),
    RemuneracaoColumnInfo("DATA_ULTIMO_STATUS", []),
    RemuneracaoColumnInfo("ULTIMO_STATUS", []),
    RemuneracaoColumnInfo("CIDADE", ["Cidade Validação"]),
    RemuneracaoColumnInfo("ESTADO", ["Estado Validação"]),
    RemuneracaoColumnInfo("AR", []),
    RemuneracaoColumnInfo("TICKET_ANTERIOR", []),
    RemuneracaoColumnInfo("PRECO", []),
    RemuneracaoColumnInfo("DATA_VENDA", ["Data de Venda", "DATA_DA_VENDA"]),
    RemuneracaoColumnInfo("PRECO BASE", []),
    RemuneracaoColumnInfo("% REM", []),
    RemuneracaoColumnInfo("R$ BIO", ["BIO"]),
    RemuneracaoColumnInfo("TOTAL R$", []),
    RemuneracaoColumnInfo("VOUCHER", []),
]


def is_remuneracao_name_matching(col_info: RemuneracaoColumnInfo, header_name: str) -> bool:
    """Verifica se um header corresponde a uma coluna de remuneração, incluindo nomes alternativos."""
    normalized_header = normalize_header(header_name)
    normalized_col_name = normalize_header(col_info.name)
    
    if normalized_header == normalized_col_name:
        return True
    
    for alt_name in col_info.alternative_names:
        normalized_alt = normalize_header(alt_name)
        if normalized_header == normalized_alt:
            return True
    
    return False


def find_data_column_for_base_column(
    base_col_name: str,
    col_info: Optional[RemuneracaoColumnInfo],
    data_header_map: dict
) -> int:
    """Encontra a coluna correspondente nos dados para uma coluna base."""
    if col_info:
        # Usa matching com nomes alternativos
        for data_header, data_col in data_header_map.items():
            if is_remuneracao_name_matching(col_info, data_header):
                return data_col
    else:
        # Usa matching simples por normalização
        normalized_base = normalize_header(base_col_name)
        for data_header, data_col in data_header_map.items():
            if normalize_header(data_header) == normalized_base:
                return data_col
    
    raise HTTPException(
        status_code=400,
        detail=f"Coluna '{base_col_name}' não encontrada na planilha enviada."
    )


@router.post(
    "/converter-remuneracao/",
    summary="Converte planilha de remuneração",
    description="Recebe uma planilha de dados (ex: Digiforte), valida suas colunas contra um template base, "
                "copia os dados e retorna o arquivo convertido."
)
async def converter_remuneracao(
    data_file: UploadFile = File(..., description="Planilha de dados a ser processada (ex: Digiforte.xlsx)")
):
    try:
        # 1. Carrega template base
        try:
            wb_base = openpyxl.load_workbook(REMUNERACAO_BASE_FILE_PATH)
            ws_base = wb_base.active
        except FileNotFoundError:
            raise HTTPException(
                status_code=500,
                detail="Erro interno no servidor: O arquivo de template base não foi encontrado."
            )

        # 2. Carrega arquivo de dados e encontra sheets relevantes
        wb_data = openpyxl.load_workbook(data_file.file)
        relevant_sheets = find_relevant_sheets(wb_data)

        # 3. Prepara mapeamentos base
        base_col_name_to_info = {col_info.name: col_info for col_info in remuneracao_columns}
        base_header_ordered = [cell.value for cell in ws_base[1] if cell.value]

        # 4. Limpa template
        if ws_base.max_row > 1:
            ws_base.delete_rows(idx=2, amount=ws_base.max_row - 1)

        # 5. Processa cada sheet e copia dados
        for ws_data in relevant_sheets:
            # Mapeia colunas para esta sheet específica
            data_header_map = {cell.value: cell.column for cell in ws_data[1] if cell.value}
            base_to_data_column_map = {}
            for base_col_name in base_header_ordered:
                if base_col_name == "VOUCHER":
                    continue
                col_info = base_col_name_to_info.get(base_col_name)
                base_to_data_column_map[base_col_name] = find_data_column_for_base_column(
                    base_col_name, col_info, data_header_map
                )

            # Processa e adiciona linhas desta sheet
            rows = process_sheet_data(ws_data, base_header_ordered, base_to_data_column_map)
            for row in rows:
                ws_base.append(row)

        # 6. Retorna arquivo convertido
        output_buffer = io.BytesIO()
        wb_base.save(output_buffer)
        output_buffer.seek(0)

        return StreamingResponse(
            output_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=Remuneracao-Convertida.xlsx"}
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"Ocorreu um erro inesperado: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Ocorreu um erro inesperado ao processar o arquivo: {str(e)}"
        )

