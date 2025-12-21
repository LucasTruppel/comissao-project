# routers/tecd.py
import openpyxl
import io
from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from starlette.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from unidecode import unidecode

from .. import crud
from ..database import engine
from ..auth import get_current_active_user
from .utils import find_resource_file

router = APIRouter(
    tags=["Conversores"],
    dependencies=[Depends(get_current_active_user)]
)

TECD_BASE_FILE_PATH = find_resource_file("Valid-Tec-D.xlsx")
TECD_PRECO_FIXO = 17


class TecdColumnInfo:
    def __init__(self, name: str, index: int, alternative_names: list[str]):
        self.name = name
        self.index = index
        self.alternative_names = alternative_names


tecd_columns = [
    TecdColumnInfo("Pedido", 1, []),
    TecdColumnInfo("Solitação", 2, ["solicitacao"]),
    TecdColumnInfo("SKU", 3, []),
    TecdColumnInfo("Produto", 4, []),
    TecdColumnInfo("Titular", 5, []),
    TecdColumnInfo("CPF", 6, []),
    TecdColumnInfo("Email", 7, []),
    TecdColumnInfo("CNPJ", 8, []),
    TecdColumnInfo("Razão Social", 9, []),
    TecdColumnInfo("Codigo Localidade", 10, []),
    TecdColumnInfo("Localidade Atendimento", 11, []),
    TecdColumnInfo("Nome Agente Validação", 12, []),
    TecdColumnInfo("CPF Agente Validação", 13, []),
    TecdColumnInfo("Data Primeira Verificação", 14, []),
    TecdColumnInfo("Código Origem", 15, []),
    TecdColumnInfo("Nome Origem", 16, []),
    TecdColumnInfo("Preço", 17, []),
    TecdColumnInfo("Serial Certificado", 18, []),
    TecdColumnInfo("Data Emissão", 19, []),
    TecdColumnInfo("Data Expiração", 20, []),
    TecdColumnInfo("Data Revogação", 21, []),
    TecdColumnInfo("Data Hora Primeira Verificação", 22, []),
    TecdColumnInfo("Data Último Status", 23, []),
    TecdColumnInfo("Último Status", 24, []),
    TecdColumnInfo("Cidade", 25, ["cidade validacao"]),
    TecdColumnInfo("Estado", 26, ["estado validacao"]),
    TecdColumnInfo("AR", 27, []),
    TecdColumnInfo("TEC-D R$ ", 28, ["preco tec-d"]),
    TecdColumnInfo("TEC-D BIO ", 29, ["bio"]),
    TecdColumnInfo("TOTAL TEC-D R$", 30, ["total tec-d"]),
]


def is_name_matching(tecd_column_info: TecdColumnInfo, header_name: str) -> bool:
    if unidecode(header_name).lower().strip() == unidecode(tecd_column_info.name).lower().strip():
        return True
    for alt_name in tecd_column_info.alternative_names:
        if unidecode(header_name).lower().strip() == unidecode(alt_name).lower().strip():
            return True
    return False


@router.post(
    "/converter-tecd/",
    summary="Converte planilha de TEC-D",
    description="Recebe uma planilha de dados, valida suas colunas contra um template base, "
                "copia os dados e retorna o arquivo convertido."
)
async def converter_tecd(
    data_file: UploadFile = File(..., description="Planilha de dados a ser processada (ex: Digiforte.xlsx)")
):
    try:
        # 1. Carrega o "template" (base)
        try:
            wb_base = openpyxl.load_workbook(TECD_BASE_FILE_PATH)
            ws_base = wb_base.active
        except FileNotFoundError:
            print(f"Erro Crítico: Arquivo base não encontrado em: {TECD_BASE_FILE_PATH}")
            raise HTTPException(
                status_code=500, 
                detail="Erro interno no servidor: O arquivo de template base não foi encontrado."
            )

        # 2. Carrega o "arquivo de dados"
        wb_data = openpyxl.load_workbook(data_file.file)
        ws_data = wb_data.active
        for sheet_name in wb_data.sheetnames:
            if unidecode(sheet_name).lower().strip() == "emissoes":
                ws_data = wb_data[sheet_name]
                break
        
        # 3. Mapeia os cabeçalhos
        column_map = {}
        for col_info in tecd_columns:
            found = False
            for cell in ws_data[1]:
                if cell.value is None:
                    continue
                if is_name_matching(col_info, cell.value):
                    column_map[col_info.index] = cell.column
                    found = True
                    break
            if not found:
                detail_msg = f"Coluna '{col_info.name}' não encontrada na planilha enviada."
                print(f"Erro de Validação: {detail_msg}")
                print(f"Colunas disponíveis: {[cell.value for cell in ws_data[1] if cell.value]}")
                raise HTTPException(status_code=400, detail=detail_msg)

        # 5. Limpa o "template"
        if ws_base.max_row > 1:
            ws_base.delete_rows(idx=2, amount=ws_base.max_row - 1)

        # 6. Copia os dados
        for data_row in ws_data.iter_rows(min_row=2):
            if all(cell.value is None for cell in data_row):
                continue
            
            new_row_values = []
            for tecd_col in tecd_columns:
                data_col_index = column_map[tecd_col.index]
                value_to_copy = data_row[data_col_index - 1].value
                new_row_values.append(value_to_copy)
            
            # Lógica específica do TEC-D
            nome_agente = new_row_values[11].strip()
            cpf_agente = new_row_values[12].replace(".", "").replace("-", "").strip()
            nome_localidade = new_row_values[10].strip()
            is_virtual = bool(nome_localidade and "VIRTUAL" in nome_localidade.upper())

            if is_virtual:
                async with AsyncSession(engine) as session:
                    db_agente = await crud.get_agente_by_cpf(session, cpf=cpf_agente)
                    if not db_agente:
                        raise HTTPException(status_code=404, detail=f"Agente {nome_agente} com CPF {cpf_agente} não encontrado.")
                    
                    db_localidade = await crud.get_localidade(session, localidade_id=db_agente.localidade_id)
                    if not db_localidade:
                        raise HTTPException(status_code=404, detail=f"Agente {nome_agente} não associado a uma localidade física.")

                    new_row_values[9] = db_localidade.codigo_localidade
                    new_row_values[10] = db_localidade.nome

            new_row_values[27] = TECD_PRECO_FIXO
            new_row_values[29] = TECD_PRECO_FIXO

            ws_base.append(new_row_values)
        
        # 7. Salva em buffer
        output_buffer = io.BytesIO()
        wb_base.save(output_buffer)
        output_buffer.seek(0)

        # 8. Retorna o buffer
        return StreamingResponse(
            output_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=Remuneracao-Convertida.xlsx"}
        )

    except HTTPException as e:
        raise e
    except Exception as e:
        print(f"Ocorreu um erro inesperado: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500, 
            detail=f"Ocorreu um erro inesperado ao processar o arquivo: {str(e)}"
        )

